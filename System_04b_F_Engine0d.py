# -*- coding: utf-8 -*-
"""
Created on Thu Jun 8 14:22:07 2023

2-Stroke Dual Fuel marine engine - Gas Mode 


@author: aahmed
"""

import cantera as ct
import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import trapz
from timeit import default_timer as timer
from scipy.interpolate import interp1d
import pandas as pd
import sys
import csv 
import os
import math
import openpyxl
from datetime import datetime

#%% 

def combustion_model(load):    
    
  
    print('''###########----||Cantera based non predictive x92DF model||----###########\n\n\t
    Cantera Version: {0}\t
    '''.format(ct.__version__), flush = True)
    
    # Suppress thermodynamic property lookup errors from Cantera
    ct.suppress_thermo_warnings()

    # Set input conditions based on the load condition
    if load == "25%":

            RPM = 55            # Engine rpm
            NE = RPM/60.             # Engine speed rps   
            wdot = 2*np.pi*NE           # angular velocity [radians/s]
            
            # Scavenge air properties
            T_scav = 273.15+37               # scav air temperature [K]
            P_scav = 0.95*10**5         # scav air pressure [bar]

            # Exhaust manif properties
            P_exh = 0.4*10**5
            T_exh = 273.15 + 300 
        
            # =============================================================================
            # Valve Timings wrt TDC = 0 
            # =============================================================================

            # Exhaust valve timing
            EVO = 115
            EVC = 220   
            exh_duration = (EVC-EVO)*1/(6*RPM)      # s/degree = 1/6*RPM
            
            # Scavenge port timing
            scav_open = 145.5
            scav_close = 214.5           
            scav_duration = (scav_close-scav_open)*1/(6*RPM)
            
            # Ignitor Injection timing
            SOI = 360-13.23 #                       
            inj_duration = 0.002 #[s]
            EOI = SOI + (inj_duration*6*RPM)   #-360??
            injection = 0.0015          # kg/cycle/cylinder, 2.079 for cylinder
            
            # Gas Admission Valve timing
            GAV_open = SOI 
            GAV_duration = 23/NE/360     # duration increased from 0.02 to 0.04 to increase residence times
            GAV_close =  GAV_open+(GAV_duration*6*RPM)
            gas_injected = (0.76/12)          # kg/cycle/cylinder
            
            # Valve flow coefficients Kv
            scav_coeff = 20
            exh_coeff = 1.5e-4
        
        

            pass
    elif load == "50%":
            # Set input conditions for 50% load
            
            RPM = 68                 # Engine rpm
            NE = RPM/60.             # Engine  rps   
            wdot = 2*np.pi*NE        # angular velocity [radians/s]
            
            # Scavenge air properties
            T_scav = 273.15+37.8               # scav air temperature [K]
            P_scav = 1.7*10**5                  # scav air pressure [bar]
            
            # Exhaust manif properties
            P_exh = 1.5*10**5
            T_exh = 273.15 + 300 
            
            # =============================================================================
            # Valve Timings 
            # =============================================================================
            
            # Exhaust valve timing
            EVO = 110
            EVC = 230                               # EVC reduced by 20 degs to maintain trapped mass 251.8    
            exh_duration = (EVC-EVO)*1/(6*RPM)      # s/degree = 1/6*RPM
            
            # Scavenge port timing
            scav_open = 145.5
            scav_close = 214.5           #changed from 251.8 for 10 def before evo
            
            scav_duration = (scav_close-scav_open)*1/(6*RPM)
            
            # Ignitor Injection timing
            SOI = 360-14
            inj_duration = 0.002 #[s]
            EOI = SOI + (inj_duration*6*RPM)   #-360??
            injection =  0.0022*1  # kg/cycle/cylinder,5% energy
            
            # Gas Admission Valve timing
            GAV_open = SOI
            GAV_duration = 25/NE/360     # duration changed to injection over 20 degs 
            GAV_close =  GAV_open+(GAV_duration*6*RPM)
            gas_injected = (1.15/12)*1         # kg/cycle/cylinder
            
            # Valve flow coefficients Kv
            scav_coeff = 1
            exh_coeff = 5e-3
            
            pass
    
    elif load== "75%":
            # Set input conditions for 75% load
            RPM = 78.                     # Engine rpm
            NE = RPM/60.
            
            # Scavenge air properties
            T_scav = 273.15+38               # scav air temperature [K]
            P_scav = 2.9*10**5         # scav air pressure [bar]
                
            # Exhaust manif properties
            P_exh = 2.7*10**5
            T_exh = 273.15 + 300 
        
            # =============================================================================
            # Valve Timings 
            # =============================================================================
            
            # Exhaust valve timing
            EVO = 106
            EVC = 230                               # exhaust close timing changed, def = 262
            exh_duration = (EVC-EVO)*1/(6*RPM)      # s/degree = 1/6*RPM
            
            # Scavenge port timing 
            scav_open = 145.5
            scav_close = 214.5           #changed from 251.8 for 10 def before evo
            scav_duration = (scav_close-scav_open)*1/(6*RPM)
                
            # Ignitor Injection timing
            SOI = 360-10
            inj_duration = 0.002                #[s]
            EOI = SOI + (inj_duration*6*RPM)    #-360??
            injection = 0.0027         # kg/cycle/cylinder, 5% total energy 
            
            # Gas Admission Valve timing
            GAV_open = SOI
            GAV_duration = 25/NE/360        # duration changed to injection over 23degs 
            GAV_close =  GAV_open+(GAV_duration*6*RPM)
            gas_injected = (1.571/12)      # kg/cycle/cylinder
            
            # Valve flow coefficients Kv
            scav_coeff = 8
            exh_coeff = 5e-3
            
            pass
    else:
            print("Invalid load condition.")
            sys.exit(1)
        
               
        # Molar Composition of air, fuel
    comp_air = 'O2:1, N2:3.7274'
    #comp_gas = 'CH4:1'
    comp_gas = 'CH4:0.952, C2H6:0.042, C3H8:0.06'   # composition from fuel analysis report
    comp_radicals = 'H:1'
        
            # fuel LHVs
    LHV = 50025.4*1000      #[J/kg]
    LHV_H2 = 119952.69*1000 #[J/kg]
    
    NE = RPM/60.                 # Engine rps   
    wdot = 2*np.pi*NE            # angular velocity [radians/s]
    
    afr = 2.7
    phi = 1/afr                 # equivalence ratio
    afr_st_gas = 17.19           # stochiometric afr for methane
            
    # Outlet/Ambient properties
    T_ambient = 300 
    P_ambient = ct.one_atm

    # Pilot injection temp and pressure: Data from 75% load 
    T_injector  = 273.15+100
    P_injector = 100*10**5          
        
    # Gas injection temp and pressure: Data from 75% load 
    T_gas = 273.15+100
    P_gas = 800*ct.one_atm   
        
    # thermal properties of intake manif
    environment = ct.Solution('air.yaml')
    cp = environment.cp_mass                         #J/kg/K
    cv = environment.cv_mass
    gamma = cp/cv
    R = ct.gas_constant/1000                        #[J/kg.K]

    #%% Save Settings and 
    # =============================================================================
    # Save settings and post calculations
    # =============================================================================
    # number of cycles to simulate, use N.5 cycles always to end at BDC and for plots to work appropriately
    sim_revs = 10.5
    # saves cylinder state and other states into a csv file
    save = 1
    # cycles from end over which to calculate results (for steady state)
    selected_cycles = 1

    #%% Mechanism Import
    mechanism = ct.Solution('gri30.yaml')

    #%% Engine Geometric Data

    B = 0.92                                                    # cylinder bore [m]
    s = 3.468                                                   # stroke length [m]
    crank_rad = 1.734                                           # crank radius [m]
    c_rod = 3.468                                               # connecting rod length [m]
    R = c_rod/crank_rad
    rc = 12.4                                                     # compression ratio                
    a_piston = np.pi*B**2*0.25
    V_swept = a_piston*s                                        # swept volume
    V_d = V_swept/(rc-1)                                        # dead/clearance volume
    dead_height = V_d/a_piston                                  # clearance volume height
    V_tot = V_swept + V_d                                       # total volume


    #%% Necessary functions

    def mdot_igniter(t):
        """Create an inlet for the injection of radicals, supplied as a Gaussian pulse."""
        total = injection   
        width = inj_duration  # width of the pulse [s]
        POI = deg_to_rad(SOI)+deg_to_rad(EOI)/2
        
        n = cycle_counter(t)
        
        t0 = (n/NE)+time(POI)  # time of fuel pulse peak [s]
        amplitude = total / (width * np.sqrt(2*np.pi))
        return amplitude * np.exp(-(t-t0)**2 / (2*width**2))

    def mdot_gas(t):
        """Create an inlet for the injection of gas, supplied as a Gaussian pulse."""
        total = gas_injected  # mass of fuel [kg], calculated for 75% load data 
        width = GAV_duration  # width of the pulse [s]
        POI = deg_to_rad(GAV_open)+deg_to_rad(GAV_close)/2
        
        n = cycle_counter(t)
        
        t0 = (n/NE)+time(POI)  # time of fuel pulse peak [s]
        amplitude = total / (width * np.sqrt(2*np.pi))
        return amplitude * np.exp(-(t-t0)**2 / (2*width**2))

    def time(CA):
        ' converts crank angle to time'
        return CA/(6*RPM)
        
    def deg_to_rad(deg):
        '''Convert degrees to radians'''
        return deg*(np.pi)/180

    def crank_angle(t):
        """Convert time to crank angle in radians """
        #np.mod(2*np.pi*RPM/60 * t + np.pi, 2*np.pi)   for when BDC is normalised to 0
        #
        return np.mod(2 * np.pi*RPM/60 * t, 2*np.pi)

    def piston_speed(t):
        """Intantaneous piston speed as a function of time/crank angle"""
        return  - s / 2 * 2 * np.pi * NE * np.sin(crank_angle(t))

    def piston_speed_inst(t):
         theta = crank_angle(t)*180/np.pi
         Vp_mean = 2*s*NE 
         Vp = Vp_mean * 0.5 * np.pi * np.sin(theta) * (1+ np.cos(theta)/R)
         return Vp

    def woschni(t):
        """Not functional: Calculates convective heat transfer acc to Woschni (1967), see Amesim help 7.6.4 Modelling incylinder process """

        # Woschni coefficients. c1 for intake/exhaust, c2 for combustion/expansion
        c11 = 6.18 
        c12 = 2.28
        c21 = 0.00324 
        c22 = 0
                
        P = cyl.thermo.P
        T = cyl.thermo.T
        Vc = cyl.volume
        P0 = 75          # pressure if combustion doesn't occur
        
        CA = crank_angle(t)
        SOI1 = SOI - 360
       
        # duration of c11 validity, from EVO to EVC
        c11_delta = np.mod(deg_to_rad(EVC-EVO), 2*np.pi)
        
        # duration of c21 validity, from start of combustion to EVO
        # note: Is SOI valid as start of combustion
        c21_delta = np.mod(deg_to_rad(EVO-SOI), 2*np.pi)


        # Setting c1 for different stages, c11 for EVO to IVC, c12 for IVC to EVO
        if np.mod( CA - deg_to_rad(EVO), 2*np.pi)<c11_delta:        #checking if current CA is less than EVC
            c1 = c11  
        else:
            c1 = c12
         
        if np.mod( CA - deg_to_rad(SOI), 2*np.pi)<c21_delta:   #checking if current CA is less than EVO
            c2 = c21
        else:
            c2 = c22
            
        # finding index of crank angle just before combustion (SOI)
        T1 = 0 
        P1 = 1 
        V1 = 1 
        for i in range(len(states.ca) - 1, -1, -1):
            if states.ca[i] < SOI1 and  states.ca[i]> SOI1-2:
                index = i
                break
              
            P1 = states.P[index]/1e5
            V1 = states.V[index]
            T1 = states.T[index] 

        
        Vp = 2*s*NE # mean piston speed. 
        
        if c2 == c21:
          A1 = c1*Vp + c2*Vc*(T1*(P-P0)/(P1*V1))
        else:
            A1 = c1*Vp

        
        h = 130 * pow(P*A1,0.8)/(T**0.53*B**0.2)
        
        return h

    def annand(t):
        'Function to calculate Annand convective heat transfer coefficient, according to Annand'
        
        A = 0.26              # convection factor [0.26-0.8]
        b = 0.7               # annand coefficient 
        nu = mechanism.viscosity/mechanism.density_mass               #nu (kinematic viscosity) = mu/rho
        lamb = mechanism.thermal_conductivity                         # [W/m/K]  
        
        Vp = 2 * s * wdot   # Average piston speed
        Re = Vp*B/nu        # Reynolds number
        Nu = A*pow(Re,b)    # Nusselt number
        
        #h = A*pow(Vp,0.7)*lamb/(pow(nu,0.7)*pow(B,0.3))
        h = Nu*lamb/B   
        return h

    def piston_position(t):
      'Function calculates instantaneous piston position, normalised to TDC'
      sin_alpha=np.sin(np.pi-wdot*t)/R
      cos_alpha=math.sqrt(1-(sin_alpha)**2)
      l=s*R/2
      rad = s/2
      X=rad+l-l*cos_alpha+rad*np.cos(wdot*t)
      return s-X 

    def ignition_delay(states, species):
        """
        Not functional: 
        This function computes the ignition delay from the occurence of the
        peak in species' concentration.
        """
        i_ign = states(species).Y.argmax()
        return states.t[i_ign]

    def cycle_counter(time):
        n = time / (1 / NE)
        return int(n)

    def port_area(t):
        'Returns effective intake port area as a function of crank angle'
        df = pd.read_csv('intake_data.csv')
        
        CA = df['CA'].tolist()  
        Area = df['Area'].tolist()
        
        current_CA = crank_angle(t)*180/np.pi    
        A = interp1d(CA, Area, kind='linear')#, fill_value='extrapolate')
        
        return float(A(current_CA))

    def pressure_function(t):
        """ Not functional: Calculates flow through intake given pressures, equation 6.13 from Heywood """
        cd=1
        A=port_area(t)
        R = 8.314
        Pin = P_scav
        Tin = T_scav
        Pout = cyl.thermo.P
        
        #assuming incompressible flow
        if Pout<=Pin:
            mdot = cd*A*pow((2*Pin)/(R*Tin),0.5)*pow(Pin-Pout,0.5) 
        else:
            mdot = 0
        
        return mdot

    #%% Reservoir and Reactor definitions

    # =============================================================================
    #   Defining reservoirs, reactors 
    # =============================================================================

    # main chamber comp and reactor set up
    mechanism.TPX = T_ambient, P_ambient, comp_air            
    mechanism.set_equivalence_ratio(phi, fuel=comp_gas, oxidizer=comp_air)
    mechanism.equilibrate('UV')                                                   # setting initial mixture to equilibrium state at constant volume and internal energy
    cyl = ct.IdealGasReactor(mechanism, name = 'Main Chamber')                    # main chamber reactor as single zone
    cyl.volume = V_d                                                              # volume initialised to TDC, dead volume  
    
    # saving solution array with cylinder states and relevant data
    states = ct.SolutionArray(
        cyl.thermo,
        extra=('t', 'ca', 'V', 'm', 'mdot_in', 'mdot_fuel', 'mdot_injector', 'mdot_out', 'dWv_dt','HRR', 'HRR1','phi','heat_loss'),
        )


    # Scavenge port containing air
    mechanism.TPX = T_scav, P_scav, comp_air 
    intake_manif = ct.Reservoir(mechanism)

    # Exhaust manifold containing air at ambient conditions 
    mechanism.TPX = T_exh, P_exh, comp_air
    exh_manif = ct.Reservoir(mechanism)

    # Injector rail containing radicals
    mechanism.TPX =  T_injector, P_injector, comp_radicals
    injector_rail = ct.Reservoir(mechanism)                                  # diesel injector using prechamber solution class

    # Natural Gas reservoir
    mechanism.TPX = T_gas, P_gas, comp_gas
    gas_tank = ct.Reservoir(mechanism)

    # environment for heat losses
    environment.TPX =  T_ambient, P_ambient, comp_air
    ambient = ct.Reservoir(environment)


    #%% SIMULATION PARAMETERS -  Reactor network and flow devices
    # =============================================================================
    #  Reactor network and simulation parameters 
    # =============================================================================

    delta_T_max = 20.
    rtol = 1.e-10           # relative error tolerance
    atol = 1.e-16           # absolute error tolerance

    # simulate with a resolution of 1/2  deg crank angle
    dt = 1./(2*360 * NE )
    t_stop = sim_revs / NE

    # Reactor network
    sim = ct.ReactorNet([cyl])
    sim.verbose = False
    sim.rtol, sim.atol = rtol, atol
    cyl.set_advance_limit('temperature', delta_T_max)

    # =============================================================================
    # Flow control devices and walls
    # =============================================================================

    # intake port from manifold to cylinder reactor
    scav_port = ct.Valve(intake_manif, cyl)
    scav_delta = np.mod(deg_to_rad(scav_close-scav_open), 2*np.pi)                  
    scav_port.valve_coeff = scav_coeff
    scav_port.set_time_function(lambda t: np.mod(crank_angle(t) - deg_to_rad(scav_open), 2*np.pi)<scav_delta)

    # Radicals  injector
    injector = ct.MassFlowController(injector_rail, cyl)
    injector_delta = np.mod(deg_to_rad(EOI-SOI), 2*np.pi)
    injector.mass_flow_coeff = injection/inj_duration
    injector.set_time_function(lambda t: np.mod(crank_angle(t) - deg_to_rad(SOI), 2*np.pi)<injector_delta)

    # GAV from gas supply to cylinder 
    gav = ct.MassFlowController(gas_tank, cyl)
    gav_delta = np.mod(deg_to_rad(GAV_close-GAV_open), 2*np.pi)
    gav_t_open = deg_to_rad(GAV_close-GAV_open)/wdot
    gav.mass_flow_coeff = gas_injected/gav_t_open 
    gav.set_time_function(lambda t: np.mod(crank_angle(t) - deg_to_rad(GAV_open), 2*np.pi)<gav_delta)

    # exhaust valve
    exh_valve = ct.Valve (cyl, exh_manif)
    exh_delta = np.mod(deg_to_rad(EVC-EVO), 2*np.pi)
    exh_valve.valve_coeff = exh_coeff
    exh_valve.set_time_function(lambda t: np.mod(crank_angle(t) - deg_to_rad(EVO), 2*np.pi)<exh_delta)



    # =============================================================================
    # Walls and Piston Heat transfer
    # =============================================================================.

    #piston wall 
    environment.TP =  450, P_ambient
    piston = ct.Wall(ambient,cyl,name='piston',K=0)
    piston.area = a_piston
    piston.set_velocity(piston_speed)
    piston.heat_transfer_coeff = annand(sim.time)

    # cylinder head
    environment.TP =  450, P_ambient #wall temp
    cyl_head = ct.Wall(cyl,ambient)
    cyl_head.area = 1
    cyl_head.heat_transfer_coeff = annand(sim.time)

    # cylinder walls
    environment.TP =  450, P_ambient
    cyl_wall = ct.Wall(cyl,ambient)
    cyl_wall.area=np.pi * B * (dead_height + piston_position(sim.time))
    cyl_wall.heat_transfer_coeff = annand(sim.time)

    # Annand radiation heat transfer #value 0.6 from AMEsim
    cyl_wall.emissivity = 0.6
    cyl_head.emissivity = 0.6
    piston.emissivity = 0.6


    #%% Simulation
    # =============================================================================
    # Run Simulation
    # =============================================================================

    # set up output data arrays
    output_data = []
    pspeed = []
    piston_X = []


    print('finished reactor setup, beginning solver...')

    fmt = '{:10.3f}  {:10.2f}  {:10.2f} {:10.2f} {:10.4f}  {:10.4g}  {:10.4g}  {:10.4g}  {:10.3g}  {:10.3g} {:10.3g}'
    print('{:10}  {:10}  {:10}  {:10}  {:10}  {:10}  {:10}   {:10}   {:10}  {:10}  {:10}  {:10}'.format(
        'time [s]', 'CA [deg]', 'P [bar]', 'T[K]',  'V [m^3]', 'phi' ,'HR [kW]', 'mass[kg]', 'mdot_a [kg/s]', 'mdot_f [kg/s]', 'mdot_e [kg/s]', 'mdot_radicals [kg/s]'))

    # =============================================================================
    #   Start simulation
    # =============================================================================
    start = timer()

    while sim.time < t_stop:
        
        # not injecting fuel in the first cycle
        if sim.time < 0.5/NE:
            gav.mass_flow_coeff=0
        else: 
            gav.mass_flow_coeff=gas_injected/gav_t_open  
        
           
        # perform time integration
        sim.advance(sim.time + dt)

        heat_transfer_coeff = annand(sim.time)  #[W/m2/K]
        
        heat_loss = heat_transfer_coeff*((np.pi * B * (dead_height + piston_position(sim.time)))+a_piston+1)*(cyl.thermo.T-450) # [W]
       
        # instantaneous power [Watts]
        dWv_dt = - (cyl.thermo.P - ambient.thermo.P) * a_piston * piston_speed(sim.time)    
        
        # calculate piston position and speed 
        piston_X.append(piston_position(sim.time))
        pspeed.append(piston_speed(sim.time))

        print(fmt.format(sim.time, crank_angle(sim.time)*180/np.pi, cyl.thermo.P/10**5, cyl.thermo.T, cyl.volume, mechanism.equivalence_ratio(),
                              np.sum(np.dot(mechanism.net_production_rates,mechanism.partial_molar_enthalpies))*cyl.volume/1000, cyl.mass, scav_port.mass_flow_rate, gav.mass_flow_rate, exh_valve.mass_flow_rate, injector.mass_flow_rate))
         
        # crank angle resolved  as -180 to 180, 0 = TDC    
        CA = crank_angle(sim.time)*180/np.pi
        if CA >= 0 and CA <= 180:
            CA = CA
        else:
            CA = CA-360
            
        # append output data to print to csv
        output_data.append((sim.time, 
                            CA,
                            cyl.thermo.P/1e5, cyl.thermo.T, cyl.volume, 
                            np.dot(mechanism.net_production_rates,mechanism.partial_molar_enthalpies)*cyl.volume/1e6,
                            dWv_dt, 
                            cyl.mass, 
                            mechanism.equivalence_ratio(), 
                            scav_port.mass_flow_rate,
                            gav.mass_flow_rate,
                            injector.mass_flow_rate,
                            exh_valve.mass_flow_rate,
                            heat_loss/1e6,
                            heat_transfer_coeff
                            # np.multiply(np.reshape(states('ch4').Y,len(states.m)),states.m)*1000
                            ))
        
        # append output data to store in Solution Array
        states.append(cyl.thermo.state,
                      t=sim.time, ca=CA,
                      V=cyl.volume, m=cyl.mass,
                      mdot_in=scav_port.mass_flow_rate,
                      mdot_fuel = gav.mass_flow_rate,
                      mdot_injector = injector.mass_flow_rate,
                      mdot_out=exh_valve.mass_flow_rate,
                      dWv_dt=dWv_dt,
                      HRR = np.dot(mechanism.net_production_rates,mechanism.partial_molar_enthalpies),
                      HRR1 = np.dot(mechanism.net_production_rates,mechanism.partial_molar_enthalpies)*cyl.volume/1000,
                      phi = mechanism.equivalence_ratio(),
                      heat_loss = heat_loss/1e6)
    end = timer()    

    ch4_mass = np.multiply(np.reshape(states('ch4').Y,len(states.m)),states.m)*1000 #[grams]

    print('Simulation time elapsed: {:.3f} seconds'.format(end - start))
    
    
    if save == 1:
        filename = "trial_DF{}.csv".format(load.replace("%", ""))
        with open(filename, 'w', newline="") as outfile:
              csvfile = csv.writer(outfile)
              csvfile.writerow(
                  ['time (s)', 'CA (deg)', 'P(bar)', 'T(K)', 'V (m3)', 'Heat Release (MW)','dWv_dt (Vmean)', 'cylinder mass (kg)', 'phi', 'mdot_in', 'mdot_fuel', 'mdot_radicals', 
                  'mdot_out', 'heat_loss [MW]', 'heat_coeff [W/m2/K]']
                  )
              csvfile.writerows(output_data)
              
    
        print('Output written to file', filename)
        print('Directory: '+os.getcwd())


    air = trapz(states.mdot_in, states.t)
    fuel = trapz(states.mdot_fuel, states.t)
    injection_tot = trapz(states.mdot_injector, states.t)
    exh = trapz(states.mdot_out, states.t)


    #%%  Plots - CHOOSE DIRECTORY TO SAVE PLOTS
    
    dpi = 600
    directory = r'C:\Users\hbusson\Documents\TNTM\PythonScripts\engine\ayub'

    def ca_ticks(t):
        """function converts time to rounded crank angle."""
        return np.floor(crank_angle(t) * 180 / np.pi/180)*180


    t = states.t
    # time interval list with each cycle time intervals, at 180 degrees
    xticks = np.arange(0, sim_revs/NE, 0.5/NE )  

    # # pressure and temperature
    fig, ax = plt.subplots(nrows=2)
    ax[0].plot(t, states.P / 1.e5)
    ax[0].set_ylabel('$p$ [bar]')
    ax[0].set_xlabel(r'$\phi$ [deg]')
    ax[0].set_xticklabels([])
    ax[1].plot(t, states.T)
    ax[1].set_ylabel('$T$ [K]')
    ax[1].set_xlabel(r'$\Crank angle$ [deg]')
    ax[1].set_xticks(xticks, xticks, rotation = 'vertical')
    ax[1].set_xticklabels(ca_ticks(xticks))
    plt.show()

    # interactive P
    plt.plot(t, states.P/1.e5)
    plt.xlabel('cycles')
    plt.ylabel('Pressure (bar)')
    plt.title('Cylinder Pressure over cycles')
    plt.gca().set(xmargin=0)  # Ensure full width of x-axis is visible
    plt.tight_layout()  # Adjust layout for better spacing
    plt.show()


    # equivalence ratio
    fig, ax = plt.subplots()
    ax.plot(t, states.phi)
    ax.set_ylabel('r$\phi$ [equivalence ratio]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.savefig(os.path.join(directory,'Equivalence Ratio - {} Load.png'.format(load)), dpi = dpi)
    plt.show()

    # cylinder mass
    fig, ax = plt.subplots()
    ax.plot(t, states.m)
    ax.set_ylabel('cylinder mass [kg]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.show()

    # gas composition (air/fuel/burned gas) by molar fractions
    air_x = (states('o2').X + states('n2').X)
    fuel_x = (states('CH4').X)
    h_x = (states('h').X)             # H radicals
    bg_x = 1-air_x-fuel_x-h_x

    fig, ax = plt.subplots()
    ax.plot(t, air_x, label='Air')
    ax.plot(t, fuel_x, label = 'Fuel')
    ax.plot(t, bg_x, label = 'Burned Gases')
    ax.legend(loc='best')
    ax.set_ylabel(' Cylinder composition [molar fractions]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.show()

    # exh gas composition by mass
    fig, ax = plt.subplots()
    ax.plot(t, np.multiply(np.reshape(states('o2').Y,len(states.m)),exh), label='Air')
    ax.plot(t, np.multiply(np.reshape(states('co2').Y,len(states.m)),exh), label='CO2')
    ax.plot(t, np.multiply(np.reshape(states('co').Y,len(states.m)),exh), label='CO')
    ax.plot(t, np.multiply(np.reshape(states('ch4').Y,len(states.m)),exh), label='CH4')
    ax.plot(t, np.multiply(np.reshape(states('H').Y,len(states.m)),exh), label='H')
    ax.plot(t, np.multiply(np.reshape(states('h2o').Y,len(states.m)),exh), label='H2O')
    #ax.plot(t, np.multiply(np.reshape(states('no').Y,len(states.m)),exh), label='NO')
    ax.legend(loc=0)
    ax.set_ylabel('Exhaust mass [kg]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.show()

    # heat of reaction and expansion work
    fig, ax = plt.subplots()
    ax.plot(t, 1.e-6 * states.heat_release_rate * states.V, label=r'$\dot{Q}$')
    ax.plot(t, 1.e-6 * states.dWv_dt, label=r'$\dot{W}_v$')
    ax.set_ylim(-1.3e2, 1e3)
    ax.legend(loc=0)
    ax.set_ylabel('Work Done and Heat Release[MW]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.show()


    # # internal energy
    # fig, ax = plt.subplots()
    # ax.plot(t, states.int_energy_mass*states.m/1e6, label=r'$\dot{U}$')
    # ax.legend(loc=0)
    # ax.set_ylabel('Internal Energy [MJ]')
    # ax.set_xlabel(r'$\phi$ [deg]')
    # ax.set_xticks(xticks, xticks, rotation = 'vertical')
    # ax.set_xticklabels(ca_ticks(xticks))
    # plt.show()

    # mass flow rates 
    fig, ax = plt.subplots()
    ax.plot (t[t > 1/NE], states.mdot_in[t > 1/NE], label = 'mdot air')
    ax.plot (t[t > 1/NE], states.mdot_out[t > 1/NE], label = 'mdot exhaust')
    ax.plot (t[t > 1/NE], states.mdot_fuel[t > 1/NE], label = 'mdot fuel')
    ax.plot (t[t > 1/NE], states.mdot_injector[t > 1/NE], label = 'radicals')
    ax.legend(loc=0)
    ax.set_ylabel('mass flow rates, (kg/s)')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.show()

    # # #piston position
    # fig, ax = plt.subplots()
    # ax.plot(t, piston_X)
    # ax.set_xlabel('t')
    # ax.set_ylabel('piston position [m]')
    # ax.set_xticks(xticks, xticks, rotation = 'vertical')
    # ax.set_xticklabels(ca_ticks(xticks))
    # plt.show()

    # #piston speed
    # for i in range(len(t)):
    #     pspeed[i] = piston_speed(t[i])
    # fig, ax = plt.subplots()
    # ax.plot(t, pspeed)
    # ax.set_xlabel('t')
    # ax.set_ylabel('piston speed m/s')
      
    
    # Species composition by molar fractions
    fig, ax = plt.subplots()
    #ax.plot(t, states('o2').X, label='O2')
    ax.plot(t, states('co2').X, label='CO2')
    ax.plot(t, states('co').X, label='CO')
    ax.plot(t, states('ch4').X, label='CH4')
    #ax.plot(t,states('h2o').X, label='H2O')
    ax.plot(t,states('no').X, label='NO')
    ax.legend(loc='best')
    ax.set_ylabel('Species conc. [molar fractions]')
    ax.set_xlabel(r'Crank Angle [deg]')
    ax.set_xticks(xticks, xticks, rotation = 'vertical')
    ax.set_title('Incylinder species concentrations')
    ax.set_xticklabels(ca_ticks(xticks))
    plt.savefig(os.path.join(directory,'cylinder composition {} Load.png'.format(load)), dpi = dpi)
    plt.show()


    # Plotting results over selected cycles for steady state
    xticks = np.arange(0, selected_cycles/NE, 0.5/NE )  
    directory = r'C:\Users\hbusson\Documents\TNTM\PythonScripts\engine\ayub'


    # Load data from csv file
    df = pd.read_csv('C:/Users/hbusson/Documents/TNTM/PythonScripts/engine/ayub/pressure real vs cantera.csv')
   
    Preal = df['P_MCR' + load.replace("%", "")].tolist()
    CAreal = df['CA_MCR' + load.replace("%", "")].tolist()

    # find time of final cycle start
    last_cycle_start = t[-1] - selected_cycles/NE
    start_index = np.argmin(np.abs(t - last_cycle_start))   #finding index of start selected no of cycles
    end_index = len(t)

    # Extract the data for selected cycles
    last_cycle_CA = states.ca[start_index:end_index ]
    last_cycle_pressure = states.P[start_index:end_index]/1e5
    Pmax = max(last_cycle_pressure) # max pressure 


    # Temperature plot of selected cycles
    last_cycle_temp = states.T[start_index:end_index]
    plt.plot(last_cycle_CA, last_cycle_temp, label = 'Temperature')
    plt.xlabel('Crank Angle (degrees)')
    plt.ylabel('Temperature [K]')
    plt.title('Temperature profile - {} Load'.format(load))
    plt.legend(loc='best')
    plt.grid(True)
    plt.show()

    # Heat loss of last cycle
    last_cycle_hloss = states.heat_loss[start_index:end_index]
    plt.plot(last_cycle_CA, last_cycle_hloss, label = 'convection factor A = 0.26')
    plt.xlabel('Crank Angle (degrees)')
    plt.ylabel(' Thermal losses [MW]')
    plt.title('Wall heat losses - {} Load, Annand heat transfer'.format(load))
    plt.legend(loc='best')
    plt.grid(True)
    plt.show()

    
     # p-V diagram
    fig, ax = plt.subplots()
    ax.plot(states.V[start_index:end_index] * 1000, states.P[start_index:end_index]/ 1.e5)
    ax.set_xlabel('$V$ [l]')
    ax.set_ylabel('$p$ [bar]')
    ax.set_title('PV diagram, {} Load'.format(load))
    # plt.savefig(os.path.join(directory,'PV {} Load.png'.format(load)), dpi = dpi)
    plt.show()


    # HRR
    fig, ax = plt.subplots()
    ax.plot(last_cycle_CA, states.heat_release_rate[start_index:end_index]/1e6*states.V[start_index:end_index], label = 'states')
    ax.set_xlabel('Crank Angle (deg)')
    ax.set_ylabel('Heat release rate (MW) ')
    ax.set_title('Heat Release Rate - {} Load'.format(load))
    # plt.savefig(os.path.join(directory,'Heat Release Rate {} Load'.format(load)), dpi = dpi)
    plt.show()

    # # work done
    fig, ax = plt.subplots()
    ax.plot(last_cycle_CA, 1.e-6 * states.dWv_dt[start_index:end_index], label=r'$\dot{W}_v$')
    ax.legend(loc=0)
    ax.set_ylabel('Power [MW]')
    ax.set_xlabel(r'$\phi$ [deg]')
    ax.set_title('Instantaneous Power - {} Load'.format(load))
    # plt.savefig(os.path.join(directory,'Instantaneous Power - {} Load'.format(load)), dpi = dpi)
    plt.show()
    
    # Plot pressure vs crank angle for selected cycles
    plt.plot(last_cycle_CA, last_cycle_pressure, label = 'Cantera DF Model')
    plt.plot(CAreal, Preal, label='Experimental data')
    plt.xlabel('Crank Angle (degrees)')
    plt.ylabel('Pressure (bar)')
    plt.title('Pressure Curve - {} Load'.format(load))
    #plt.savefig(os.path.join(directory,'Pressure {} Load.png'.format(load)), dpi = dpi)
    plt.legend(loc='best')
    plt.grid(True)
    plt.show()


    


    #%% Post Calculations

    output_str = '{:45s}{:>4.1f} {}'

    # =============================================================================
    # Calculations over whole simulation duration
    # =============================================================================

    # heat release [kW]

    # energy released
    Q = trapz(states.heat_release_rate * states.V, t)               # [J]

    print(output_str.format('Heat release rate per cylinder (estimate):',
                            Q / t[-1] / 1000., 'kW'))


    # # Combustion efficiency
    ## Note: I am calculating this wrong, or there seems to be an imbalance of energy
    # fuel_energy = (LHV*fuel)+(LHV_H2*injection_tot)    #[J]
    # energy_released = Q
    # eta_comb = energy_released/fuel_energy
    # print('Combustion efficiency: {:.2f}'.format(eta_comb*100), '%')


    # expansion power
    W = trapz(states.dWv_dt, t)
    print(output_str.format('Expansion power per cylinder (estimate):',
                            W / t[-1] / 1000., 'kW'))
    print(output_str.format('Engine power (Indicated):',
                            W*12 / t[-1] / 1000., 'kW'))

    # efficiency
    eta = W / Q
    print(output_str.format('Efficiency (estimate):', eta * 100., '%'))


    ######################################################################
    # Results over selected cycles - Steady state
    ######################################################################


    print ('\n Calculating results over selected cycles (for steady state) \n')

    air = trapz(states.mdot_in[start_index:end_index], states.t[start_index:end_index])
    fuel = trapz(states.mdot_fuel[start_index:end_index], states.t[start_index:end_index])
    injection_tot = trapz(states.mdot_injector[start_index:end_index], states.t[start_index:end_index])
    exh = trapz(states.mdot_out[start_index:end_index], states.t[start_index:end_index])

    # heat release of selected cycles [kW]
    Q = trapz(states.heat_release_rate[start_index:end_index]* states.V[start_index:end_index], t[start_index:end_index])     # energy released [J]
    print(output_str.format('Heat release rate per cylinder:', 
                            Q / (t[-1]-t[start_index]) / 1000., 'kW'))
    print(output_str.format('Heat release rate (engine):', 
                            Q*12 / (t[-1]-t[start_index]) / 1000., 'kW'))
    # Thermal losses
    heat_loss_energy = trapz(states.heat_loss[start_index:end_index]*1e6, t[start_index:end_index])
    print(output_str.format('Thermal losses :',heat_loss_energy*12/(t[-1]-t[start_index])/1000, 'kW'))
    
    print(output_str.format('Pmax (comb):', 
                            Pmax, 'bars'))

    # Calculating Power [kW] 
    cyl_power = trapz (states.dWv_dt[start_index:end_index],t[start_index:end_index])\
                        /(t[-1]-t[start_index])              #cylinder power [W]
    IMEP = (cyl_power)/V_swept/1e5          
    FMEP = 0.3695 + 4.7853e-05*RPM + IMEP*2.9877e-02        # emperical correlation for FMEP 
    BMEP = IMEP - FMEP 
    brake_pow = BMEP*1e5*V_swept*12/1000
    fric_pow = FMEP*1e5*V_swept*12/1000

    print ('\n Power')

    print(output_str.format('Engine power (Indicated):',cyl_power*12/1000, 'kW'))
    print(output_str.format('Engine power (Friction):',fric_pow, 'kW'))
    print(output_str.format('Engine power (Brake):',brake_pow, 'kW'))
    
    # efficiency
    eta = cyl_power*12 /(Q*12 / (t[-1]-t[start_index]))
    print(output_str.format('Efficiency :', eta * 100., '%'))
    
    print ('\n Mean Effective Pressures')
    print(output_str.format('IMEP: ',IMEP, 'bar'))
    print(output_str.format('FMEP: ',FMEP, 'bar'))
    print(output_str.format('BMEP: ',BMEP, 'bar'))
    
    print ('\n Inlet & Exhaust pressures')
    print(output_str.format('Inlet pressure: ',P_scav/1e5, 'bar'))
    print(output_str.format('Exhaust receiver pressure: ',P_exh/1e5, 'bar'))


    
    print ('\n Fuel data')
    # Fuel composition
    print('Natural Gas composition (molar):             ', comp_gas)
    
    #Equivalence ratio
    print('states',states)
    print('phi',type(phi))
   # print(output_str.format('Equivalence Ratio :', max(states,phi)))
    
 

    
    # Engine fuel consumption 
    fuel_cons = (fuel/selected_cycles)*RPM*60*12  #[kg/h], full engine
    print(output_str.format('Fuel consumption (gas):', fuel_cons, 'kg/h'))

    # bsfc over selected cycles
    brake_work = brake_pow*1000*(t[-1]-t[start_index])/12
    bsfc = fuel*1000/(brake_work/(3.6*1e6))
    bsfc_pilot = injection_tot*1000/(brake_work/(3.6*1e6))
    print(output_str.format('BSFC (gas):', bsfc, 'g/kWh'))
    print(output_str.format('BSFC (pilot):', bsfc_pilot, 'g/kWh'))



    # =============================================================================
    # Emissions
    # =============================================================================
    print ('\n Emissions ')

    # CO emissions
    MW = states.mean_molecular_weight
    CO_emission = trapz(MW * states.mdot_out * states('CO').X[:, 0], t)
    CO_emission /= trapz(MW * states.mdot_out, t)
    print(output_str.format('CO emission (estimate):', CO_emission * 1.e6, 'ppm'))

    # CO2 emissions
    MW = states.mean_molecular_weight
    CO2_emission = trapz(MW * states.mdot_out * states('CO2').X[:, 0], t)
    CO2_emission /= trapz(MW * states.mdot_out, t)
    print(output_str.format('CO2 emission (estimate):', CO2_emission * 1.e6, 'ppm'))

    # NO emissions
    MW = states.mean_molecular_weight
    NO_emission = trapz(MW * states.mdot_out * states('NO').X[:, 0], t)
    NO_emission /= trapz(MW * states.mdot_out, t)
    print(output_str.format('NO emission (estimate):', NO_emission * 1.e6, 'ppm'))

    # CH4 emissions
    MW = states.mean_molecular_weight
    CH4_emission = trapz(MW * states.mdot_out * states('CH4').X[:, 0], t)
    CH4_emission /= trapz(MW * states.mdot_out, t)
    print(output_str.format('CH4 slip (estimate):', CH4_emission * 1.e6, 'ppm'))
    print(output_str.format('CH4 slip % (estimate):', CH4_emission*100/fuel, '%'))



    #%% Results saved with Input parameters 

    if save == 1:
        
        
        filename = 'Results and Parameters, v0.5.xlsx'
        sheetname = load
        
        avg_fuel_per_cycle=fuel/selected_cycles
        parameters = ['SOI', 'EOI', 'inj_duration', 'injection', 'GAV_open', 'GAV_duration', 'GAV_close', 'gas_injected', 'avg_fuel_per_cycle',
                      'P_scav/1e5', 'T_scav', 'exh_manif.thermo.P/1e5', 'scav_port.valve_coeff', 'exh_valve.valve_coeff', 'EVC', 'Pmax', 'brake_pow', 'bsfc', 'IMEP', 'BMEP', 'sim_revs', 'selected_cycles']
        
   
        # Load or create the Excel file
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        
        # Check if the sheet already exists
        if load in workbook.sheetnames:
            sheet = workbook[load]
        else:
            sheet = workbook.create_sheet(load)
        
        next_row = sheet.max_row + 1
        
        # Write the data to the sheet
        sheet.cell(row=next_row, column=1, value=f'Case {next_row}')
        for i, param in enumerate(parameters, start=2):
            value = eval(param) if i <= len(parameters) else None
            sheet.cell(row=next_row, column=i, value=value)
        
        # Save the updated workbook to the Excel file
        workbook.save(filename)
        # # Save pressure plot to folder
            
        # folder = r'C:\Users\aahmed\Pictures\v1.0'
        # file_name = f'{next_row}.png'
        # dpi = 600
        # plt.plot(last_cycle_CA, last_cycle_pressure, label = 'Cantera Model')
        # plt.plot(CAreal, Preal, label='Experimental data')
        # plt.xlabel('Crank Angle (degrees)')
        # plt.ylabel('Pressure')
        # plt.title('Pressure Curve - {} Load'.format(load))
        # plt.legend(loc='best')
        # plt.grid(True)
        # os.makedirs(folder, exist_ok=True)
        # plt.savefig(os.path.join(folder, file_name), dpi=dpi)

#%% Main 
        
def main():
    # load selection
    print("Select the load condition:")
    print("1. 25%")
    print("2. 50%")
    print("3. 75%")
    choice = "1"

    if choice == "1":
        load = "25%"
    elif choice == "2":
        load = "50%"
    elif choice == "3":
        load = "75%"
    else:
        print("Invalid choice.")
        sys.exit(1) 
    combustion_model(load)


# Run the program
if __name__ == "__main__":
    main()